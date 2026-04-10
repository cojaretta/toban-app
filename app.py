from flask import Flask, request, send_file, render_template_string
import math, io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

app = Flask(__name__)

HTML = """
<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>給食・そうじ当番表メーカー</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: 'Hiragino Maru Gothic ProN', 'BIZ UDPGothic', sans-serif;
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    min-height: 100vh;
    display: flex;
    align-items: center;
    justify-content: center;
    padding: 20px;
  }
  .card {
    background: white;
    border-radius: 20px;
    padding: 40px;
    width: 100%;
    max-width: 480px;
    box-shadow: 0 20px 60px rgba(0,0,0,0.2);
  }
  .emoji { font-size: 48px; text-align: center; margin-bottom: 8px; }
  h1 {
    text-align: center;
    font-size: 22px;
    color: #333;
    margin-bottom: 6px;
  }
  .subtitle {
    text-align: center;
    font-size: 13px;
    color: #888;
    margin-bottom: 32px;
  }
  .form-group { margin-bottom: 20px; }
  label {
    display: block;
    font-size: 13px;
    font-weight: bold;
    color: #555;
    margin-bottom: 6px;
  }
  input {
    width: 100%;
    padding: 12px 16px;
    border: 2px solid #e0e0e0;
    border-radius: 10px;
    font-size: 16px;
    font-family: inherit;
    transition: border-color 0.2s;
    outline: none;
  }
  input:focus { border-color: #667eea; }
  .row { display: flex; gap: 12px; }
  .row .form-group { flex: 1; }
  .hint {
    font-size: 11px;
    color: #aaa;
    margin-top: 4px;
  }
  button {
    width: 100%;
    padding: 16px;
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    color: white;
    border: none;
    border-radius: 12px;
    font-size: 17px;
    font-family: inherit;
    font-weight: bold;
    cursor: pointer;
    margin-top: 8px;
    transition: opacity 0.2s, transform 0.1s;
  }
  button:hover { opacity: 0.9; transform: translateY(-1px); }
  button:active { transform: translateY(0); }
  .note {
    margin-top: 20px;
    padding: 14px;
    background: #f8f9ff;
    border-radius: 10px;
    font-size: 12px;
    color: #666;
    line-height: 1.7;
  }
  .note strong { color: #667eea; }
  {% if error %}
  .error {
    background: #fff0f0;
    border: 1px solid #ffcccc;
    border-radius: 10px;
    padding: 12px 16px;
    color: #cc0000;
    font-size: 13px;
    margin-bottom: 20px;
  }
  {% endif %}
</style>
</head>
<body>
<div class="card">
  <div class="emoji">🍱</div>
  <h1>給食・そうじ当番表メーカー</h1>
  <p class="subtitle">クラス情報を入力してExcelをダウンロード</p>

  {% if error %}
  <div class="error">⚠️ {{ error }}</div>
  {% endif %}

  <form method="POST" action="/generate">
    <div class="form-group">
      <label>📋 クラス名</label>
      <input type="text" name="class_name" placeholder="例：3年2組"
             value="{{ class_name or '' }}" required>
    </div>

    <div class="row">
      <div class="form-group">
        <label>👥 人数</label>
        <input type="number" name="num_students" placeholder="32"
               value="{{ num_students or '32' }}" min="10" max="45" required>
        <p class="hint">10〜45人</p>
      </div>
      <div class="form-group">
        <label>📅 開始週</label>
        <input type="number" name="start_week" placeholder="1"
               value="{{ start_week or '1' }}" min="1" max="52" required>
        <p class="hint">途中追加は週番号を指定</p>
      </div>
    </div>

    <div class="form-group">
      <label>🗓️ 週数</label>
      <input type="number" name="num_weeks" placeholder="35"
             value="{{ num_weeks or '35' }}" min="1" max="52" required>
      <p class="hint">通常35週・追加の場合は必要な週数だけ</p>
    </div>

    <button type="submit">⬇️ Excelをダウンロード</button>
  </form>

  <div class="note">
    <strong>ダウンロード後にExcelで入力するもの</strong><br>
    ・係名シート：給食当番・そうじ場所・仕事内容を入力（全週に自動反映）<br>
    ・名前シート：名前を入力
  </div>
</div>
</body>
</html>
"""

def make_excel(class_name, num, start_week, num_weeks):
    HALF  = math.ceil(num / 2)
    HALF2 = num - HALF

    # モノクロ配色
    C_BLACK  = "000000"
    C_DARK   = "333333"   # タイトル背景
    C_MID    = "666666"   # ヘッダー背景
    C_LGRAY  = "CCCCCC"   # ヘッダー2段目背景
    C_STRIPE = "F0F0F0"   # データ行ストライプ
    C_WHITE  = "FFFFFF"
    C_BORDER = "888888"

    thin  = Side(style="thin",   color=C_BORDER)
    med   = Side(style="medium", color=C_BLACK)
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    bdr_t = Border(left=med, right=med, top=med, bottom=thin)

    def mf(sz=10, bold=False, color="000000"):
        return Font(name="Meiryo", size=sz, bold=bold, color=color)
    def fl(c): return PatternFill("solid", start_color=c)
    def put(ws, row, col, val="", bg=None, ft=None, ha="center", va="center"):
        c = ws.cell(row=row, column=col, value=val)
        if bg: c.fill = fl(bg)
        if ft: c.font = ft
        c.alignment = Alignment(horizontal=ha, vertical=va)
        c.border = bdr
        return c

    wb = Workbook()
    wb.remove(wb.active)

    # 青系グラデーション配色
    C_TITLE  = "1A3A5C"   # 濃紺（タイトル）
    C_HDR    = "2E6DA4"   # 中青（列ヘッダー）
    C_SUBHDR = "D6E8F7"   # 薄青（説明行）
    C_ROW0   = "EBF4FB"   # 行背景 偶数
    C_ROW1   = "FFFFFF"   # 行背景 奇数
    C_ACCENT = "1A5276"   # アクセント文字

    # ── 係名シート ─────────────────────────────────
    ws_ky = wb.create_sheet("係名")
    ws_ky.sheet_view.showGridLines = False
    ws_ky.column_dimensions["A"].width = 14   # 給食当番
    ws_ky.column_dimensions["B"].width = 14   # そうじ場所
    ws_ky.column_dimensions["C"].width = 16   # そうじ仕事内容
    ws_ky.column_dimensions["D"].width = 16   # ①ヘッダー名
    ws_ky.column_dimensions["E"].width = 16   # ②ヘッダー名
    ws_ky.column_dimensions["F"].width = 30   # ヒント

    # タイトル行
    ws_ky.merge_cells("A1:E1")
    c = ws_ky["A1"]
    c.value = f"🍱  係名入力シート　― {class_name} ―"
    c.font  = mf(12, True, "FFFFFF")
    c.fill  = fl(C_TITLE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws_ky.row_dimensions[1].height = 30

    # 説明行
    ws_ky.merge_cells("A2:E2")
    c = ws_ky["A2"]
    c.value = "★ ここに入力すると当番表の全週に自動反映されます"
    c.font  = mf(9, color=C_ACCENT)
    c.fill  = fl(C_SUBHDR)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws_ky.row_dimensions[2].height = 18

    # 列ヘッダー
    ky_headers = ["給食当番", "そうじ場所", "そうじ仕事内容", "①列のヘッダー名", "②列のヘッダー名"]
    for ci, label in enumerate(ky_headers):
        c = ws_ky.cell(row=3, column=ci+1, value=label)
        c.font = mf(10, True, "FFFFFF")
        c.fill = fl(C_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr
    ws_ky.row_dimensions[3].height = 22

    # ヒント列
    c = ws_ky.cell(row=3, column=6, value="← ①②列のヘッダーを変えたい場合に入力（例：給食当番おやすみ）")
    c.font = mf(8, color="888888")
    c.alignment = Alignment(horizontal="left", vertical="center")

    # D4=①デフォルト、E4=②デフォルト
    for ci, val in enumerate(["①", "②"], start=4):
        c = ws_ky.cell(row=4, column=ci, value=val)
        c.font = mf(10, color=C_ACCENT)
        c.fill = fl(C_ROW0)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = bdr
    ws_ky.row_dimensions[4].height = 18

    # データ行（給食・掃除係、HALF行分）
    for i in range(HALF):
        r = 4 + i
        ws_ky.row_dimensions[r].height = 18
        row_bg = C_ROW0 if i % 2 == 0 else C_ROW1
        for ci in range(3):   # A〜C列
            c = ws_ky.cell(row=r, column=ci+1, value="")
            c.fill = fl(row_bg)
            c.font = mf(10)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = bdr
        # D・E列（①②ヘッダー）は4行目のみ入力欄
        if r > 4:
            for ci in [4, 5]:
                c = ws_ky.cell(row=r, column=ci, value="")
                c.fill = fl(C_ROW1)
                c.border = bdr

    # ── 名前シート ─────────────────────────────────
    ws_nm = wb.create_sheet("名前")
    ws_nm.sheet_view.showGridLines = False
    ws_nm.column_dimensions["A"].width = 5
    ws_nm.column_dimensions["B"].width = 14

    # タイトル行
    ws_nm.merge_cells("A1:B1")
    c = ws_nm["A1"]
    c.value = f"👤  名前入力　― {class_name}・{num}人 ―"
    c.font  = mf(11, True, "FFFFFF")
    c.fill  = fl(C_TITLE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws_nm.row_dimensions[1].height = 30

    # 説明行
    ws_nm.merge_cells("A2:B2")
    c = ws_nm["A2"]
    c.value = "★ 名前を入力してください"
    c.font  = mf(9, color=C_ACCENT)
    c.fill  = fl(C_SUBHDR)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws_nm.row_dimensions[2].height = 16

    # 列ヘッダー
    put(ws_nm, 3, 1, "番号", bg=C_HDR, ft=mf(10, True, "FFFFFF"))
    put(ws_nm, 3, 2, "名前", bg=C_HDR, ft=mf(10, True, "FFFFFF"))
    ws_nm.row_dimensions[3].height = 20

    # 名前行
    for n in range(1, num + 1):
        r = 3 + n
        ws_nm.row_dimensions[r].height = 18
        row_bg = C_ROW0 if n % 2 == 0 else C_ROW1
        put(ws_nm, r, 1, n,  bg=row_bg, ft=mf(10, bold=True, color=C_ACCENT))
        put(ws_nm, r, 2, "", bg=row_bg, ft=mf(11))

    # ── 当番表シート ─────────────────────────────
    ws = wb.create_sheet("当番表")
    ws.sheet_view.showGridLines = False

    C_KY=2; C_SJ1=3; C_SJ2=4; C_NM1=5; C_NM2=6

    ws.column_dimensions["A"].width = 0.8
    ws.column_dimensions["B"].width = 10.5
    ws.column_dimensions["C"].width = 10.5
    ws.column_dimensions["D"].width = 10.5
    ws.column_dimensions["E"].width = 12.0
    ws.column_dimensions["F"].width = 12.0
    ws.column_dimensions["G"].width = 0.8

    ws.page_setup.paperSize   = 9
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left   = 0.6 / 2.54
    ws.page_margins.right  = 0.6 / 2.54
    ws.page_margins.top    = 0.8 / 2.54
    ws.page_margins.bottom = 0.8 / 2.54

    NM_RANGE  = f"名前!$B$4:$B${3 + num}"
    KY_RANGE  = f"係名!$A$4:$A${3+HALF}"
    SJ1_RANGE = f"係名!$B$4:$B${3+HALF}"
    SJ2_RANGE = f"係名!$C$4:$C${3+HALF}"

    # 罫線定義
    # t=thin(内側), m=medium(外枠・区切り)
    T = Side(style="thin",   color=C_BORDER)
    M = Side(style="medium", color=C_BLACK)

    def cell_bdr(left=T, right=T, top=T, bottom=T):
        return Border(left=left, right=right, top=top, bottom=bottom)

    # セクション区切り縦線（給食|そうじ|名前の間）は medium
    # 通常セルの上下左右は thin
    # 外枠は medium

    def set_row(ws, row, col, val, bg, font, left=T, right=T, top=T, bottom=T):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = fl(bg)
        c.font      = font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = Border(left=left, right=right, top=top, bottom=bottom)
        return c

    ROWS_PER_WEEK = 1 + 2 + HALF + 2

    for wi, week in enumerate(range(start_week, start_week + num_weeks)):
        base    = wi * ROWS_PER_WEEK + 1
        r_title = base
        r_hdr1  = base + 1
        r_hdr2  = base + 2
        r_data0 = base + 3
        r_sep1  = base + 3 + HALF
        r_sep2  = base + 3 + HALF + 1

        if wi > 0:
            ws.row_breaks.append(Break(id=base))

        ws.row_dimensions[r_title].height = 28
        ws.row_dimensions[r_hdr1].height  = 15
        ws.row_dimensions[r_hdr2].height  = 19
        for i in range(HALF):
            ws.row_dimensions[r_data0 + i].height = 21
        ws.row_dimensions[r_sep1].height = 5
        ws.row_dimensions[r_sep2].height = 5

        # ── 週タイトル（全列結合・外枠medium）──
        ws.merge_cells(start_row=r_title, start_column=C_KY,
                       end_row=r_title,   end_column=C_NM2)
        c = ws.cell(row=r_title, column=C_KY,
                    value=f"第 {week} 週　　{class_name}　給食・そうじ当番")
        c.font      = mf(13, True, C_BLACK)
        c.fill      = fl(C_WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = Border(left=M, right=M, top=M, bottom=M)
        for col in range(C_KY+1, C_NM2):
            ws.cell(row=r_title, column=col).border = Border(top=M, bottom=M)
        ws.cell(row=r_title, column=C_NM2).border = Border(right=M, top=M, bottom=M)

        # ── ヘッダー1段目（グループ名）──
        # 給食当番（単独）
        set_row(ws, r_hdr1, C_KY, "給食当番", C_WHITE, mf(9, True, C_BLACK),
                left=M, right=M, top=T, bottom=M)

        # そうじ当番（C_SJ1〜C_SJ2 結合）
        ws.merge_cells(start_row=r_hdr1, start_column=C_SJ1,
                       end_row=r_hdr1,   end_column=C_SJ2)
        c = ws.cell(row=r_hdr1, column=C_SJ1, value="そうじ当番")
        c.font = mf(9, True, C_BLACK); c.fill = fl(C_WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(left=M, right=T, top=T, bottom=M)
        ws.cell(row=r_hdr1, column=C_SJ2).border = Border(left=T, right=M, top=T, bottom=M)

        # 名前ヘッダー（C_NM1〜C_NM2）は表示しない・罫線だけ設定
        ws.merge_cells(start_row=r_hdr1, start_column=C_NM1,
                       end_row=r_hdr1,   end_column=C_NM2)
        c = ws.cell(row=r_hdr1, column=C_NM1, value="")
        c.font = mf(9); c.fill = fl(C_WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(left=M, right=T, top=T, bottom=M)
        ws.cell(row=r_hdr1, column=C_NM2).border = Border(left=T, right=M, top=T, bottom=M)

        # ── ヘッダー2段目（列名）──
        set_row(ws, r_hdr2, C_KY,  "係",       C_WHITE, mf(9, True, C_BLACK),
                left=M, right=M, top=T, bottom=M)
        set_row(ws, r_hdr2, C_SJ1, "場所",     C_WHITE, mf(9, True, C_BLACK),
                left=M, right=T, top=T, bottom=M)
        set_row(ws, r_hdr2, C_SJ2, "仕事内容", C_WHITE, mf(9, True, C_BLACK),
                left=T, right=M, top=T, bottom=M)
        set_row(ws, r_hdr2, C_NM1, '=IFERROR(係名!$D$4,"①")', C_WHITE, mf(9, True, C_BLACK),
                left=M, right=T, top=T, bottom=M)
        set_row(ws, r_hdr2, C_NM2, '=IFERROR(係名!$E$4,"②")', C_WHITE, mf(9, True, C_BLACK),
                left=T, right=M, top=T, bottom=M)

        # ── データ行 ──
        shift = week - 1
        for i in range(HALF):
            row    = r_data0 + i
            row_bg = C_STRIPE if i % 2 == 0 else C_WHITE
            is_last = (i == HALF - 1)
            bot = M if is_last else T   # 最終行の下は太線

            # 給食係
            c = ws.cell(row=row, column=C_KY,
                        value=f"=IFERROR(INDEX({KY_RANGE},{i+1}),\"\")")
            c.font = mf(10); c.fill = fl(row_bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(left=M, right=M, top=T, bottom=bot)

            # そうじ場所
            c = ws.cell(row=row, column=C_SJ1,
                        value=f"=IFERROR(INDEX({SJ1_RANGE},{i+1}),\"\")")
            c.font = mf(10); c.fill = fl(row_bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(left=M, right=T, top=T, bottom=bot)

            # そうじ仕事
            c = ws.cell(row=row, column=C_SJ2,
                        value=f"=IFERROR(INDEX({SJ2_RANGE},{i+1}),\"\")")
            c.font = mf(10); c.fill = fl(row_bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(left=T, right=M, top=T, bottom=bot)

            # 名前①
            f1 = f"=IFERROR(INDEX({NM_RANGE},MOD({shift}+{i},{num})+1),\"\")"
            c = ws.cell(row=row, column=C_NM1, value=f1)
            c.font = mf(11); c.fill = fl(row_bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(left=M, right=T, top=T, bottom=bot)

            # 名前②
            f2 = (f"=IFERROR(INDEX({NM_RANGE},MOD({shift}+{HALF}+{i},{num})+1),\"\")"
                  if i < HALF2 else "")
            c = ws.cell(row=row, column=C_NM2, value=f2)
            c.font = mf(11); c.fill = fl(row_bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(left=T, right=M, top=T, bottom=bot)

        # 区切り行
        for r in [r_sep1, r_sep2]:
            for ci in range(C_KY, C_NM2 + 1):
                ws.cell(row=r, column=ci).fill = fl("E8E8E8")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.route("/")
def index():
    return render_template_string(HTML)


@app.route("/generate", methods=["POST"])
def generate():
    try:
        class_name   = request.form["class_name"].strip()
        num_students = int(request.form["num_students"])
        start_week   = int(request.form["start_week"])
        num_weeks    = int(request.form["num_weeks"])

        if not class_name:
            raise ValueError("クラス名を入力してください")
        if not (10 <= num_students <= 45):
            raise ValueError("人数は10〜45人で入力してください")
        if not (1 <= start_week <= 52):
            raise ValueError("開始週は1〜52で入力してください")
        if not (1 <= num_weeks <= 52):
            raise ValueError("週数は1〜52で入力してください")

        buf = make_excel(class_name, num_students, start_week, num_weeks)
        safe_name = class_name.replace(" ", "_").replace("　", "_")
        filename  = f"{safe_name}_当番表_第{start_week}週〜{num_weeks}週分.xlsx"

        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except ValueError as e:
        return render_template_string(HTML, error=str(e),
            class_name=request.form.get("class_name"),
            num_students=request.form.get("num_students"),
            start_week=request.form.get("start_week"),
            num_weeks=request.form.get("num_weeks"))


if __name__ == "__main__":
    app.run(debug=True)
